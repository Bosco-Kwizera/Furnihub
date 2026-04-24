from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Sum, Q
from django.utils import timezone
from datetime import datetime, timedelta
from apps.products.models import Product, Category, ProductImage
from apps.orders.models import Order, OrderStatusHistory
from django.contrib.auth.models import User, Group, Permission
from django.http import JsonResponse, HttpResponse
from decimal import Decimal
import csv
from openpyxl import Workbook


def admin_required(view_func):
    """Decorator to check if user is admin/superuser"""
    actual_decorator = user_passes_test(
        lambda u: u.is_authenticated and u.is_superuser,
        login_url='/accounts/login/'
    )
    return actual_decorator(view_func)


@admin_required
def dashboard(request):
    """Admin dashboard home"""
    # Get statistics
    total_orders = Order.objects.count()
    total_products = Product.objects.count()
    total_users = User.objects.count()
    
    # FIXED: Calculate total revenue from ALL orders (not just paid)
    # This ensures you see revenue even if payment status is not updated
    total_revenue = Order.objects.aggregate(Sum('total'))['total__sum'] or Decimal('0')
    
    # For detailed breakdown
    paid_revenue = Order.objects.filter(payment_status='paid').aggregate(Sum('total'))['total__sum'] or Decimal('0')
    pending_revenue = Order.objects.filter(payment_status='pending').aggregate(Sum('total'))['total__sum'] or Decimal('0')
    cash_on_delivery_revenue = Order.objects.filter(payment_method='cash_on_delivery').aggregate(Sum('total'))['total__sum'] or Decimal('0')
    mobile_money_revenue = Order.objects.filter(payment_method='mobile_money').aggregate(Sum('total'))['total__sum'] or Decimal('0')
    
    # Recent orders
    recent_orders = Order.objects.order_by('-created_at')[:5]
    
    # Low stock products
    low_stock = Product.objects.filter(stock_quantity__lt=10, is_active=True)[:5]
    
    # Orders by status
    orders_by_status = {
        'pending': Order.objects.filter(status='pending').count(),
        'processing': Order.objects.filter(status='processing').count(),
        'shipped': Order.objects.filter(status='shipped').count(),
        'delivered': Order.objects.filter(status='delivered').count(),
        'cancelled': Order.objects.filter(status='cancelled').count(),
    }
    
    # Orders by payment status
    payment_by_status = {
        'paid': Order.objects.filter(payment_status='paid').count(),
        'pending': Order.objects.filter(payment_status='pending').count(),
        'failed': Order.objects.filter(payment_status='failed').count(),
        'refunded': Order.objects.filter(payment_status='refunded').count(),
    }
    
    # Orders by payment method
    payment_by_method = {
        'cash_on_delivery': Order.objects.filter(payment_method='cash_on_delivery').count(),
        'mobile_money': Order.objects.filter(payment_method='mobile_money').count(),
        'paypal': Order.objects.filter(payment_method='paypal').count(),
        'stripe': Order.objects.filter(payment_method='stripe').count(),
    }
    
    # Recent activity
    recent_activity = []
    
    # Get recent orders
    for order in Order.objects.order_by('-created_at')[:10]:
        recent_activity.append({
            'type': 'order',
            'message': f'New order #{order.order_number} - ${order.total}',
            'time': order.created_at,
            'order_id': order.id
        })
    
    # Get today's stats
    today = timezone.now().date()
    today_orders = Order.objects.filter(created_at__date=today).count()
    today_revenue = Order.objects.filter(created_at__date=today).aggregate(Sum('total'))['total__sum'] or Decimal('0')
    
    context = {
        'total_orders': total_orders,
        'total_products': total_products,
        'total_users': total_users,
        'total_revenue': total_revenue,
        'paid_revenue': paid_revenue,
        'pending_revenue': pending_revenue,
        'cash_on_delivery_revenue': cash_on_delivery_revenue,
        'mobile_money_revenue': mobile_money_revenue,
        'today_orders': today_orders,
        'today_revenue': today_revenue,
        'recent_orders': recent_orders,
        'low_stock': low_stock,
        'orders_by_status': orders_by_status,
        'payment_by_status': payment_by_status,
        'payment_by_method': payment_by_method,
        'recent_activity': recent_activity[:10],
    }
    return render(request, 'admin_dashboard/dashboard.html', context)


@admin_required
def orders_list(request):
    """List all orders"""
    status_filter = request.GET.get('status', '')
    orders = Order.objects.all().order_by('-created_at')
    
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    context = {
        'orders': orders,
        'status_filter': status_filter,
        'status_choices': Order.STATUS_CHOICES,
    }
    return render(request, 'admin_dashboard/orders_list.html', context)


@admin_required
def order_detail(request, order_id):
    """View order details"""
    order = get_object_or_404(Order, id=order_id)
    context = {'order': order}
    return render(request, 'admin_dashboard/order_detail.html', context)


@admin_required
def update_order_status(request, order_id):
    """Update order status"""
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        new_status = request.POST.get('status')
        note = request.POST.get('note', '')
        
        old_status = order.status
        order.status = new_status
        order.save()
        
        # Add to status history
        OrderStatusHistory.objects.create(
            order=order,
            status=new_status,
            note=f'Status changed from {old_status} to {new_status}. {note}',
            created_by=request.user
        )
        
        messages.success(request, f'Order #{order.order_number} status updated to {order.get_status_display()}')
    return redirect('admin_dashboard:order_detail', order_id=order_id)


# ==================== PAYMENT STATUS UPDATE VIEW ====================

@admin_required
def update_payment_status(request, order_id):
    """Update payment status of an order"""
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        new_payment_status = request.POST.get('payment_status')
        note = request.POST.get('note', '')
        
        old_status = order.payment_status
        
        # Update payment status
        order.payment_status = new_payment_status
        order.save()
        
        # Add to status history
        OrderStatusHistory.objects.create(
            order=order,
            status=order.status,
            note=f'Payment status changed from {old_status} to {new_payment_status}. {note}',
            created_by=request.user
        )
        
        messages.success(request, f'Payment status for Order #{order.order_number} updated to {new_payment_status}')
        
    return redirect('admin_dashboard:order_detail', order_id=order_id)


@admin_required
def products_list(request):
    """List all products"""
    products = Product.objects.all().order_by('-created_at')
    context = {'products': products}
    return render(request, 'admin_dashboard/products_list.html', context)


@admin_required
def add_product(request):
    """Add new product"""
    if request.method == 'POST':
        # Get form data
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        price = request.POST.get('price')
        stock_quantity = request.POST.get('stock_quantity')
        description = request.POST.get('description')
        short_description = request.POST.get('short_description', '')
        brand = request.POST.get('brand', '')
        material = request.POST.get('material', '')
        color = request.POST.get('color', '')
        dimensions = request.POST.get('dimensions', '')
        is_active = request.POST.get('is_active') == 'on'
        is_featured = request.POST.get('is_featured') == 'on'
        
        # Create product
        product = Product.objects.create(
            name=name,
            category_id=category_id,
            price=price,
            stock_quantity=stock_quantity,
            description=description,
            short_description=short_description,
            brand=brand,
            material=material,
            color=color,
            dimensions=dimensions,
            is_active=is_active,
            is_featured=is_featured
        )
        
        # Handle images
        images = request.FILES.getlist('images')
        for i, image in enumerate(images):
            ProductImage.objects.create(
                product=product,
                image=image,
                is_primary=(i == 0),
                order=i
            )
        
        messages.success(request, f'Product "{product.name}" added successfully!')
        return redirect('admin_dashboard:products_list')
    
    categories = Category.objects.all()
    context = {'categories': categories}
    return render(request, 'admin_dashboard/product_form.html', context)


@admin_required
def edit_product(request, product_id):
    """Edit product"""
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        product.name = request.POST.get('name')
        product.category_id = request.POST.get('category')
        product.price = request.POST.get('price')
        product.stock_quantity = request.POST.get('stock_quantity')
        product.description = request.POST.get('description')
        product.short_description = request.POST.get('short_description', '')
        product.brand = request.POST.get('brand', '')
        product.material = request.POST.get('material', '')
        product.color = request.POST.get('color', '')
        product.dimensions = request.POST.get('dimensions', '')
        product.is_active = request.POST.get('is_active') == 'on'
        product.is_featured = request.POST.get('is_featured') == 'on'
        product.save()
        
        # Handle new images
        images = request.FILES.getlist('images')
        for image in images:
            ProductImage.objects.create(
                product=product,
                image=image,
                is_primary=False
            )
        
        messages.success(request, f'Product "{product.name}" updated successfully!')
        return redirect('admin_dashboard:products_list')
    
    categories = Category.objects.all()
    context = {'product': product, 'categories': categories}
    return render(request, 'admin_dashboard/product_form.html', context)


@admin_required
def delete_product(request, product_id):
    """Delete product"""
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product_name = product.name
        product.delete()
        messages.success(request, f'Product "{product_name}" deleted successfully!')
    return redirect('admin_dashboard:products_list')


@admin_required
def categories_list(request):
    """List all categories"""
    categories = Category.objects.all()
    context = {'categories': categories}
    return render(request, 'admin_dashboard/categories_list.html', context)


@admin_required
def add_category(request):
    """Add new category"""
    if request.method == 'POST':
        name = request.POST.get('name')
        parent_id = request.POST.get('parent')
        description = request.POST.get('description')
        
        Category.objects.create(
            name=name,
            parent_id=parent_id if parent_id else None,
            description=description
        )
        messages.success(request, f'Category "{name}" added successfully!')
        return redirect('admin_dashboard:categories_list')
    
    categories = Category.objects.all()
    context = {'categories': categories}
    return render(request, 'admin_dashboard/category_form.html', context)


# ==================== USER MANAGEMENT VIEWS ====================

@admin_required
def users_list(request):
    """List all users with options to manage roles"""
    users = User.objects.all().order_by('-date_joined')
    
    # Get all available groups/roles
    groups = Group.objects.all()
    
    context = {
        'users': users,
        'groups': groups,
    }
    return render(request, 'admin_dashboard/users_list.html', context)


@admin_required
def user_detail(request, user_id):
    """View and edit user details"""
    user_obj = get_object_or_404(User, id=user_id)
    groups = Group.objects.all()
    
    if request.method == 'POST':
        # Update basic info
        user_obj.first_name = request.POST.get('first_name', '')
        user_obj.last_name = request.POST.get('last_name', '')
        user_obj.email = request.POST.get('email', '')
        user_obj.is_active = request.POST.get('is_active') == 'on'
        user_obj.is_staff = request.POST.get('is_staff') == 'on'
        user_obj.is_superuser = request.POST.get('is_superuser') == 'on'
        user_obj.save()
        
        # Update groups
        user_obj.groups.clear()
        group_ids = request.POST.getlist('groups')
        for group_id in group_ids:
            group = Group.objects.get(id=group_id)
            user_obj.groups.add(group)
        
        messages.success(request, f'User "{user_obj.username}" updated successfully!')
        return redirect('admin_dashboard:users_list')
    
    context = {
        'user_obj': user_obj,
        'groups': groups,
        'user_groups': user_obj.groups.all(),
    }
    return render(request, 'admin_dashboard/user_detail.html', context)


@admin_required
def user_roles(request):
    """Manage user roles and groups"""
    groups = Group.objects.all()
    
    if request.method == 'POST':
        # Create new group
        if 'create_group' in request.POST:
            group_name = request.POST.get('group_name')
            if group_name:
                group, created = Group.objects.get_or_create(name=group_name)
                if created:
                    messages.success(request, f'Group "{group_name}" created successfully!')
                else:
                    messages.error(request, f'Group "{group_name}" already exists!')
        
        # Delete group
        elif 'delete_group' in request.POST:
            group_id = request.POST.get('group_id')
            group = get_object_or_404(Group, id=group_id)
            group_name = group.name
            group.delete()
            messages.success(request, f'Group "{group_name}" deleted successfully!')
        
        # Update group permissions
        elif 'update_permissions' in request.POST:
            group_id = request.POST.get('group_id')
            group = get_object_or_404(Group, id=group_id)
            permission_ids = request.POST.getlist('permissions')
            group.permissions.clear()
            for perm_id in permission_ids:
                perm = Permission.objects.get(id=perm_id)
                group.permissions.add(perm)
            messages.success(request, f'Permissions for "{group.name}" updated successfully!')
        
        return redirect('admin_dashboard:user_roles')
    
    # Get all permissions grouped by app
    permissions_by_app = {}
    for perm in Permission.objects.all().order_by('content_type__app_label', 'codename'):
        app_label = perm.content_type.app_label
        if app_label not in permissions_by_app:
            permissions_by_app[app_label] = []
        permissions_by_app[app_label].append({
            'id': perm.id,
            'name': perm.name,
            'codename': perm.codename
        })
    
    context = {
        'groups': groups,
        'permissions_by_app': permissions_by_app,
    }
    return render(request, 'admin_dashboard/user_roles.html', context)


@admin_required
def toggle_user_status(request, user_id):
    """Activate/Deactivate user account"""
    user_obj = get_object_or_404(User, id=user_id)
    user_obj.is_active = not user_obj.is_active
    user_obj.save()
    
    status = "activated" if user_obj.is_active else "deactivated"
    messages.success(request, f'User "{user_obj.username}" has been {status}.')
    return redirect('admin_dashboard:users_list')


@admin_required
def make_staff(request, user_id):
    """Make a user staff member"""
    user_obj = get_object_or_404(User, id=user_id)
    user_obj.is_staff = True
    user_obj.save()
    messages.success(request, f'User "{user_obj.username}" is now a staff member.')
    return redirect('admin_dashboard:users_list')


@admin_required
def make_superuser(request, user_id):
    """Make a user superuser"""
    user_obj = get_object_or_404(User, id=user_id)
    user_obj.is_superuser = True
    user_obj.is_staff = True
    user_obj.save()
    messages.success(request, f'User "{user_obj.username}" is now a superuser.')
    return redirect('admin_dashboard:users_list')


@admin_required
def remove_staff(request, user_id):
    """Remove staff status from user"""
    user_obj = get_object_or_404(User, id=user_id)
    user_obj.is_staff = False
    user_obj.is_superuser = False
    user_obj.save()
    messages.success(request, f'Staff status removed from "{user_obj.username}".')
    return redirect('admin_dashboard:users_list')


@admin_required
def get_group_permissions(request, group_id):
    """API endpoint to get permissions for a group"""
    group = get_object_or_404(Group, id=group_id)
    permissions = list(group.permissions.values_list('id', flat=True))
    return JsonResponse({'permissions': permissions})


# ==================== REPORTS & ANALYTICS VIEWS ====================

@admin_required
def reports_view(request):
    """Main reports dashboard"""
    # Get date filters
    date_range = request.GET.get('date_range', 'today')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Calculate date range based on selection
    now = timezone.now()
    today = now.date()
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        date_range = 'custom'
    elif date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = today - timedelta(days=1)
    elif date_range == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif date_range == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    elif date_range == 'year':
        start_date = today - timedelta(days=365)
        end_date = today
    else:
        start_date = today
        end_date = today
    
    # Filter orders by date range
    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).order_by('-created_at')
    
    # Calculate statistics
    total_orders = orders.count()
    total_revenue = orders.aggregate(Sum('total'))['total__sum'] or Decimal('0')
    avg_order_value = total_revenue / total_orders if total_orders > 0 else Decimal('0')
    
    # Orders by status
    orders_by_status = {}
    for status_code, status_label in Order.STATUS_CHOICES:
        orders_by_status[status_label] = orders.filter(status=status_code).count()
    
    # Daily breakdown
    daily_data = []
    current_date = start_date
    while current_date <= end_date:
        day_orders = orders.filter(created_at__date=current_date)
        daily_data.append({
            'date': current_date,
            'orders': day_orders.count(),
            'revenue': day_orders.aggregate(Sum('total'))['total__sum'] or Decimal('0')
        })
        current_date += timedelta(days=1)
    
    context = {
        'orders': orders,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'avg_order_value': avg_order_value,
        'orders_by_status': orders_by_status,
        'daily_data': daily_data,
        'start_date': start_date,
        'end_date': end_date,
        'date_range': date_range,
        'status_choices': Order.STATUS_CHOICES,
    }
    return render(request, 'admin_dashboard/reports.html', context)


@admin_required
def export_orders_csv(request):
    """Export orders to CSV format"""
    # Get date range from request
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        orders = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).order_by('-created_at')
    else:
        orders = Order.objects.all().order_by('-created_at')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="orders_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow([
        'Order Number', 'Customer', 'Email', 'Date', 'Status', 
        'Payment Status', 'Subtotal', 'Tax', 'Shipping', 'Total', 
        'Items Count', 'Shipping Name', 'Shipping Phone', 'Shipping Address'
    ])
    
    # Write data rows
    for order in orders:
        writer.writerow([
            order.order_number,
            order.user.username,
            order.user.email,
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.get_status_display(),
            order.payment_status,
            float(order.subtotal),
            float(order.tax),
            float(order.shipping_cost),
            float(order.total),
            order.get_items_count(),
            order.shipping_name,
            order.shipping_phone,
            f"{order.shipping_address.address_line1}, {order.shipping_address.city}, {order.shipping_address.state} {order.shipping_address.postal_code}, {order.shipping_address.country}"
        ])
    
    return response


@admin_required
def export_orders_excel(request):
    """Export orders to Excel format"""
    # Get date range from request
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        orders = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).order_by('-created_at')
    else:
        orders = Order.objects.all().order_by('-created_at')
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders Report"
    
    # Define headers
    headers = [
        'Order Number', 'Customer', 'Email', 'Date', 'Status', 
        'Payment Status', 'Subtotal', 'Tax', 'Shipping', 'Total', 
        'Items Count', 'Shipping Name', 'Shipping Phone', 'Shipping Address'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data rows
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.order_number)
        ws.cell(row=row, column=2, value=order.user.username)
        ws.cell(row=row, column=3, value=order.user.email)
        ws.cell(row=row, column=4, value=order.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=5, value=order.get_status_display())
        ws.cell(row=row, column=6, value=order.payment_status)
        ws.cell(row=row, column=7, value=float(order.subtotal))
        ws.cell(row=row, column=8, value=float(order.tax))
        ws.cell(row=row, column=9, value=float(order.shipping_cost))
        ws.cell(row=row, column=10, value=float(order.total))
        ws.cell(row=row, column=11, value=order.get_items_count())
        ws.cell(row=row, column=12, value=order.shipping_name)
        ws.cell(row=row, column=13, value=order.shipping_phone)
        ws.cell(row=row, column=14, value=f"{order.shipping_address.address_line1}, {order.shipping_address.city}, {order.shipping_address.state} {order.shipping_address.postal_code}")
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="orders_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@admin_required
def export_products_csv(request):
    """Export products to CSV format"""
    products = Product.objects.all()
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="products_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow([
        'ID', 'Name', 'Category', 'Price', 'Compare Price', 'Stock', 
        'Brand', 'Material', 'Color', 'Dimensions', 'Status', 'Featured', 'Created Date'
    ])
    
    # Write data rows
    for product in products:
        writer.writerow([
            product.id,
            product.name,
            product.category.name,
            float(product.price),
            float(product.compare_price) if product.compare_price else '',
            product.stock_quantity,
            product.brand or '',
            product.material or '',
            product.color or '',
            product.dimensions or '',
            'Active' if product.is_active else 'Inactive',
            'Yes' if product.is_featured else 'No',
            product.created_at.strftime('%Y-%m-%d')
        ])
    
    return response